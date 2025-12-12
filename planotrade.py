#!/usr/bin/env python3
# planotrade.py
import sys
import pandas as pd
import unicodedata
from openpyxl import load_workbook

# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------
def normalize_header(name: str) -> str:
    if not isinstance(name, str):
        return ''
    s = unicodedata.normalize('NFD', name.strip().upper())
    return ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')

# --------------------------------------------------------------------
# Carrega plano de trade da aba PLANOTRADE
# --------------------------------------------------------------------
def load_trade_plan(path: str) -> dict:
    # Abre planilha com data_only para ler valores de fórmulas
    wb    = load_workbook(path, data_only=True)
    sheet = next((n for n in wb.sheetnames if normalize_header(n) == 'PLANOTRADE'),
                 wb.sheetnames[0])
    ws    = wb[sheet]

    # Lê valor de CARTEIRA (célula abaixo do cabeçalho)
    carteira_val = 0.0
    for cell in ws[1]:
        if normalize_header(cell.value) == 'CARTEIRA':
            carteira_val = cell.offset(row=1).value or 0.0
            break
    wb.close()

    # Usa pandas para facilitar leitura de colunas
    df = pd.read_excel(path, sheet_name=sheet, engine='openpyxl')
    df.columns = [normalize_header(c) for c in df.columns]

    # Converte BANCA para float
    df['BANCA'] = df['BANCA'].apply(
        lambda x: float(str(x).replace('R$', '').replace('.', '').replace(',', '.'))
    )

    # Ordena e seleciona linha ativa
    df = df.sort_values('BANCA')
    mask = df['BANCA'] <= carteira_val
    row  = df[mask].iloc[-1] if mask.any() else df.iloc[0]
    row  = row.fillna(0)

    # Extrai campos da linha
    return {
        'sheet':      sheet,
        'carteira':   float(carteira_val),
        'banca':      float(row['BANCA']),
        'contratos':  int(row.get('CONTRATOS', 0)),
        'operacoes':  int(row.get('OPERACOES', 0)),
        'stop_op':    float(row.get('STOP OPERACAO', 0)),
        'meta_op':    float(row.get('META OPERACAO', 0)),
        'stop_day':   float(row.get('STOP DIARIO', 0)),
        'meta_day':   float(row.get('META DIARIA', 0))
    }

# --------------------------------------------------------------------
# Desenha painel de plano de trade no terminal
# --------------------------------------------------------------------
def draw_panel(plan: dict, start_row: int = 1):
    """
    Desenha o painel de Plano de Trade com título, cabeçalho e duas linhas: PLANO com valores e OPERACAO em branco.
    """
    labels = ('TRADE','CARTEIRA','BANCA','CONTRATOS','OPERACOES',
              'STOP OP','META OP','STOP DIARIO','META DIARIA')
    width  = 12
    hdr    = '|' + '|'.join(f"{l:^{width}}" for l in labels) + '|'
    sep    = '-' * len(hdr)

    # Título acima do painel
    sys.stdout.write(f"\x1b[{start_row};0HPLANO DE TRADE")
    # Linha de separação
    sys.stdout.write(f"\x1b[{start_row+1};0H{sep}")
    # Cabeçalho
    sys.stdout.write(f"\x1b[{start_row+2};0H{hdr}")
    # Linha de separação
    sys.stdout.write(f"\x1b[{start_row+3};0H{sep}")

    # Formatação de valores em BRL (ex.: 1.234,56)
    def fmt_brl(v):
        s = f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return s.center(width)

    # Valores da linha PLANO
    vals = [
        fmt_brl(plan['carteira']),
        fmt_brl(plan['banca']),
        str(plan['contratos']).center(width),
        str(plan['operacoes']).center(width),
        fmt_brl(plan['stop_op']),
        fmt_brl(plan['meta_op']),
        fmt_brl(plan['stop_day']),
        fmt_brl(plan['meta_day'])
    ]
    row_plano = '|' + '|'.join([
        'PLANO'.center(width),
        *vals
    ]) + '|'
    # Linha PLANO
    sys.stdout.write(f"\x1b[{start_row+4};0H{row_plano}")

    # Linha OPERACAO em branco
    blanks = [''.center(width)] * (len(labels)-1)
    row_oper  = '|' + '|'.join([
        'OPERACAO'.center(width),
        *blanks
    ]) + '|'
    sys.stdout.write(f"\x1b[{start_row+5};0H{row_oper}")

    # Linha de separação final
    sys.stdout.write(f"\x1b[{start_row+6};0H{sep}")
    sys.stdout.flush()

# --------------------------------------------------------------------
# Execução isolada
# --------------------------------------------------------------------
if __name__ == '__main__':
    print('PLANO DE TRADE')
    cfg = sys.argv[1] if len(sys.argv) > 1 else 'config.ini'
    try:
        import configparser, os
        c = configparser.ConfigParser()
        c.read(cfg)
        path = c.get('ARQUIVOS', 'FALCAO_EXCEL_PATH', fallback=c.get('DEFAULT', 'FALCAO_EXCEL_PATH'))
        path = os.path.expanduser(os.path.expandvars(path))
    except Exception as e:
        print(f"Erro lendo config.ini: {e}")
        sys.exit(1)

    plan = load_trade_plan(path)
    print(f"Planilha:  {path}")
    print(f"CARTEIRA:  {plan['carteira']:,.2f}".replace(',', 'X').replace('.', ',').replace('X','.'))
    draw_panel(plan, start_row=4)
