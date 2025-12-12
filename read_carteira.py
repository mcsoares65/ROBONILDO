#!/usr/bin/env python3
"""
read_carteira.py

Imprime no stdout o valor da coluna CARTEIRA da aba PLANOTRADE
da sua planilha configurada em config.ini.
"""

import os
import sys
import unicodedata
import configparser
from openpyxl import load_workbook

def normalize_header(name: str) -> str:
    """Remove acentos e espaços, deixa em MAIÚSCULAS."""
    if not isinstance(name, str):
        return ''
    s = unicodedata.normalize('NFD', name.strip().upper())
    return ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')

def get_excel_path(cfg_path: str = 'config.ini') -> str:
    """
    Lê config.ini para obter FALCAO_EXCEL_PATH (seção ARQUIVOS ou DEFAULT).
    Tenta UTF-8 e, em caso de erro, latin-1.
    """
    cfg = configparser.ConfigParser()
    # tenta utf-8
    try:
        cfg.read(cfg_path, encoding='utf-8')
    except Exception:
        # fallback latin1
        cfg.read(cfg_path, encoding='latin-1')

    # prioridade: seção ARQUIVOS, depois DEFAULT
    if cfg.has_section('ARQUIVOS') and cfg.has_option('ARQUIVOS', 'FALCAO_EXCEL_PATH'):
        return cfg.get('ARQUIVOS', 'FALCAO_EXCEL_PATH')
    return cfg.get('DEFAULT', 'FALCAO_EXCEL_PATH')

def read_carteira_value(excel_path: str) -> float:
    """
    Abre o arquivo Excel (somente valores) e retorna o valor em CARTEIRA.
    """
    wb = load_workbook(excel_path, data_only=True)
    # encontra aba PLANOTRADE (ou pega a primeira)
    sheet = next(
        (n for n in wb.sheetnames if normalize_header(n) == 'PLANOTRADE'),
        wb.sheetnames[0]
    )
    ws = wb[sheet]

    # encontra coluna CARTEIRA na primeira linha
    carteira_col = None
    for cell in ws[1]:
        if normalize_header(cell.value) == 'CARTEIRA':
            carteira_col = cell.col_idx
            break

    if carteira_col is None:
        wb.close()
        raise ValueError("Coluna 'CARTEIRA' não encontrada na aba PLANOTRADE.")

    # lê valor da linha 2, mesma coluna
    val = ws.cell(row=2, column=carteira_col).value
    wb.close()
    return float(val or 0.0)

def main():
    # pega o caminho da planilha
    try:
        path = get_excel_path()
        path = os.path.expanduser(os.path.expandvars(path))
    except Exception as e:
        print(f"Erro ao ler config.ini: {e}", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(path):
        print(f"Arquivo não encontrado: {path}", file=sys.stderr)
        sys.exit(1)

    try:
        carteira = read_carteira_value(path)
        # formata com vírgula e duas casas decimais
        out = f"{carteira:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        print(out)
    except Exception as e:
        print(f"Erro ao ler valor de CARTEIRA: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
