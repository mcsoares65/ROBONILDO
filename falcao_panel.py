# falcao_panel.py
import sys
import unicodedata
import pandas as pd
import colorama
from colorama import Fore, Style
from indicador_medias import COL_W, PERIODOS
from openpyxl import load_workbook

# Inicializa o Colorama para cores no terminal
colorama.init(autoreset=True)

# Parâmetros de cálculo de estocástico (mesmos do estocastico_lento)
STO_LEN = 8
D_LEN   = 3
RULES   = {tf.strip().upper(): rule for tf, rule in PERIODOS.items()}
TFS     = list(RULES.keys())


def normalize_header(name: str) -> str:
    if not isinstance(name, str):
        return ''
    s = unicodedata.normalize('NFD', name.strip().upper())
    return ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')


def load_scenarios(path: str) -> pd.DataFrame:
    # Lê aba CENARIOS da planilha com fórmulas avaliadas (data_only)
    wb    = load_workbook(path, read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if normalize_header(n) == 'CENARIOS'), wb.sheetnames[0])
    ws    = wb[sheet]
    rows  = list(ws.values)
    wb.close()
    headers = [normalize_header(h) for h in rows[0]]
    data    = rows[1:]
    return pd.DataFrame(data, columns=headers).fillna('')


def find_matching_scenario(df: pd.DataFrame, trends: dict) -> pd.Series:
    # Retorna primeira linha onde cada coluna TF bate com trends[TF]
    for _, row in df.iterrows():
        if all(
            normalize_header(str(row.get(tf, ''))) == normalize_header(trends.get(tf, ''))
            for tf in TFS
        ):
            return row
    return df.iloc[0] if not df.empty else pd.Series()


def color_val(text: str) -> str:
    # Colore DECISAO (COMPRA/VENDA) e estados de tendência
    s = (text or '').strip().upper().center(COL_W)
    val = s.strip()
    if val in ('COMPRA', 'ALTA', 'SOBRECOMPRADO'):
        return Fore.GREEN + s + Style.RESET_ALL
    if val in ('VENDA', 'BAIXA', 'SOBREVENDIDO'):
        return Fore.RED + s + Style.RESET_ALL
    if val in ('LATERAL', 'AGUARDA'):
        return Fore.YELLOW + s + Style.RESET_ALL
    return s


def draw_layout(start_row: int = 1):
    # Desenha cabeçalho estático do painel Falcão
    top    = start_row + 7
    title = (
        f"{'FALCAO'.ljust(COL_W)}|"
        f"{'CENARIO ATUAL':^{COL_W}}|"
        f"{'EM OPERACAO':^{COL_W}}|"
        f"{'GESTAO LUCRO':^{COL_W}}|"
    )
    sep = '-' * len(title)
    sys.stdout.write(f"\x1b[{top};0H{title}")
    sys.stdout.write(f"\x1b[{top+1};0H{sep}")
    for j, lbl in enumerate(('CODIGO','DECISAO','SCORE','PRECO','BREAKEVEN','TRAILING'), start=2):
        row   = top + j
        empty = ' '.center(COL_W)
        sys.stdout.write(
            f"\x1b[{row};0H{lbl.ljust(COL_W)}|{empty}|{empty}|{empty}|"
        )
    sys.stdout.write(f"\x1b[{top+8};0H{sep}")


def draw_values(df: pd.DataFrame, excel_path: str, start_row: int = 1):
    # Recalcula tendências via estocástico lento
    last_dt = pd.to_datetime(df['DataHora'].iat[-1])
    last_pr = df['Último'].iat[-1]
    serie   = df.set_index('DataHora')['Último']
    trends  = {}

    for tf in TFS:
        bars = serie.resample(RULES[tf], closed='right', label='right').ohlc().ffill()
        ts = last_dt.floor(RULES[tf])
        if ts in bars.index:
            bars.at[ts,'high']  = max(bars.at[ts,'high'], last_pr)
            bars.at[ts,'low']   = min(bars.at[ts,'low'],  last_pr)
            bars.at[ts,'close'] = last_pr
        if len(bars) >= STO_LEN:
            hr = bars['high'].rolling(STO_LEN).max()
            lr = bars['low'].rolling(STO_LEN).min()
            fk = 100 * (bars['close'] - lr) / (hr - lr)
            sk = fk.ewm(span=D_LEN).mean()
            sd = sk.ewm(span=D_LEN).mean()
            k_val, d_val = sk.iat[-1], sd.iat[-1]
        else:
            k_val, d_val = 0.0, 0.0
        if abs(k_val - d_val) < 1:
            trends[tf] = 'LATERAL'
        elif k_val > 80:
            trends[tf] = 'SOBRECOMPRADO'
        elif k_val < 20:
            trends[tf] = 'SOBREVENDIDO'
        else:
            trends[tf] = 'ALTA' if k_val > d_val else 'BAIXA'

    # Escolhe o cenário que casa com as tendências
    sce = find_matching_scenario(load_scenarios(excel_path), trends)
    top = start_row + 7
    b1  = COL_W + 1

    # Valores a exibir: seis linhas
    vals = [
        str(sce.get('CODIGO', '')),
        sce.get('DECISAO', ''),
        str(sce.get('SCORE', '')),
        f"{last_pr:.0f}",
        f"{float(sce.get('BREAKEVEN') or 0):.0f}",
        f"{float(sce.get('TRAILING') or 0):.0f}"
    ]
    for i, v in enumerate(vals, start=2):
        # CORREÇÃO: DECISAO está na segunda posição (i==3)
        if i == 3:
            text = color_val(v)
        else:
            text = v.center(COL_W)
        sys.stdout.write(f"\x1b[{top+i};{b1}H{text}")
    sys.stdout.flush()
