#!/usr/bin/env python3
# gestor_trade.py

import sys
import os
from datetime import datetime
import config
from persistencia_operacoes import init_log, append_operacao, read_operacoes
from planotrade import load_trade_plan, normalize_header
from indicador_medias import COL_W
from colorama import init as colorama_init, Fore, Style
from executa_ordem import executar_compra, executar_venda
from openpyxl import load_workbook

# configura colorama
colorama_init(autoreset=True)

# formata valor em BRL (centralizado)
def fmt_brl(v):
    s = f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return s.center(COL_W)

# lê valor do ponto índice
diff_raw = config._cfg.get('TRADE', 'VALOR_PONTO_INDICE')
VALOR_PONTO_INDICE = float(diff_raw.split('#',1)[0].strip())

# estado global
defaults = {
    'open_position': False,
    'entry_code': None,
    'entry_decision': None,
    'entry_score': 0,
    'entry_price': 0.0,
    'entry_breakeven_val': 0.0,
    'entry_trailing_val': 0.0,
    'current_price': 0.0,
    'current_score': 0
}
globals().update(defaults)
plan_data = {}
plan_values = {}
log_file_path = None
last_record = None
operations_executed = 0
daily_pnl = 0.0
max_ops = 0
contracts = 1


def init_gestor(excel_path: str):
    """
    Carrega plano de trade, inicializa arquivo de log diário e recupera histórico.
    """
    global plan_data, plan_values, log_file_path, last_record
    global operations_executed, daily_pnl, max_ops, contracts

    # carrega plano de trade
    plan = load_trade_plan(excel_path)
    plan_data = plan.copy()
    plan_values = {
        'stop_op':  plan['stop_op'],
        'meta_op':  plan['meta_op'],
        'stop_day': plan['stop_day'],
        'meta_day': plan['meta_day']
    }

    # limites diários
    contracts = plan_data.get('contratos', 1)
    max_ops = plan_data.get('operacoes', 1)
    operations_executed = 0
    daily_pnl = 0.0

    # inicializa log e lê histórico
    log_file_path = init_log()
    history = read_operacoes(log_file_path)
    last_record = history[-1] if history else None


def check_exits():
    """
    Verifica critérios de saída e, se atingidos, fecha posição,
    registra no CSV e atualiza planilha.
    """
    global open_position, operations_executed, daily_pnl, last_record

    if not open_position:
        return

    # calcula lucro/prejuízo por contrato e diário
    diff_price = (current_price - entry_price) if entry_decision == 'COMPRA' else (entry_price - current_price)
    pl_per_contract = diff_price * VALOR_PONTO_INDICE
    new_daily = daily_pnl + pl_per_contract

    # critérios de saída
    hit_op = (pl_per_contract >= plan_values['meta_op']) or (pl_per_contract <= plan_values['stop_op'])
    hit_day = (new_daily >= plan_values['meta_day']) or (new_daily <= plan_values['stop_day'])

    if hit_op or hit_day:
        # fecha posição em todos contratos
        for _ in range(contracts):
            if entry_decision == 'COMPRA':
                executar_venda()
            else:
                executar_compra()
        open_position = False

        # atualiza contagem e acumulado
        operations_executed += 1
        daily_pnl = new_daily
        op_num = operations_executed

        # prepara registro
        registro = {
            'hora':        datetime.now().strftime('%H:%M:%S'),
            'op_num':      op_num,
            'carteira':    plan_data['carteira'] + daily_pnl,
            'banca':       daily_pnl,
            'contratos':   contracts,
            'operacoes':   max_ops,
            'stop_op':     plan_values['stop_op'],
            'meta_op':     plan_values['meta_op'],
            'stop_diario': plan_values['stop_day'],
            'meta_diaria': plan_values['meta_day']
        }
        append_operacao(log_file_path, registro)
        last_record = registro

        # atualiza célula resumo na planilha (linha 2, coluna CARTEIRA)
        try:
            path = os.path.expanduser(os.path.expandvars(config.FALCAO_EXCEL_PATH))
            wb = load_workbook(path)
            ws = wb[plan_data.get('sheet')] if plan_data.get('sheet') in wb.sheetnames else wb.active
            col_idx = None
            for cell in ws[1]:
                if normalize_header(cell.value) == 'CARTEIRA':
                    col_idx = cell.col_idx
                    break
            if col_idx:
                cell = ws.cell(row=2, column=col_idx)
                cell.value = registro['carteira']
                cell.number_format = '#,##0.00'
                wb.save(path)
        except Exception as e:
            print(f"Erro salvando Excel: {e}")


def update_trade_state(scenario, last_price: float):
    """
    Atualiza preço/score, tenta saídas e abre nova entrada conforme limites.
    """
    global open_position, entry_code, entry_decision, entry_score, entry_price
    global entry_breakeven_val, entry_trailing_val, current_price, current_score

    current_price = last_price
    current_score = int(scenario.get('SCORE') or 0)
    dec = str(scenario.get('DECISAO','')).strip().upper()

    # checa stop/meta
    check_exits()
    if open_position:
        return

    # nova entrada
    if operations_executed < max_ops and dec in ('COMPRA','VENDA'):
        for _ in range(contracts):
            if dec == 'COMPRA':
                executar_compra()
            else:
                executar_venda()
        open_position = True
        entry_code = scenario.get('CODIGO')
        entry_decision = dec
        entry_score = current_score
        entry_price = current_price
        entry_breakeven_val = float(scenario.get('BREAKEVEN') or 0)
        entry_trailing_val = float(scenario.get('TRAILING') or 0)


def draw_operacao(start_row: int = 1):
    """
    Desenha painel FALCAO (EM OPERAÇÃO / GESTAO LUCRO)
    e linha OPERACAO do PLANO DE TRADE.
    """
    top = start_row + 7
    b1 = COL_W + 1
    b2 = b1 + (COL_W + 1)
    b3 = b2 + (COL_W + 1)

    # limpa painel FALCAO
    for i in range(6):
        row = top + 2 + i
        sys.stdout.write(f"\x1b[{row};{b2}H{'':^{COL_W}}")
        sys.stdout.write(f"\x1b[{row};{b3}H{'':^{COL_W}}")
    sys.stdout.flush()

    # desenha FALCAO se houver posição
    if open_position:
        op_vals = [
            str(entry_code),
            entry_decision,
            str(entry_score),
            f"{entry_price:.0f}",
            f"{entry_breakeven_val:.0f}",
            f"{entry_trailing_val:.0f}"
        ]
        diff_score = current_score - entry_score
        diff_price = (current_price - entry_price) if entry_decision=='COMPRA' else (entry_price - current_price)
        valor_operacao = diff_price * VALOR_PONTO_INDICE
        brk_price = entry_price + (entry_breakeven_val if entry_decision=='COMPRA' else -entry_breakeven_val)
        trl_price = entry_price + (entry_trailing_val if entry_decision=='COMPRA' else -entry_trailing_val)

        # formata diff_score
        if diff_score>0:
            diff_score_str = Fore.GREEN + str(diff_score).center(COL_W) + Style.RESET_ALL
        elif diff_score<0:
            diff_score_str = Fore.RED + str(diff_score).center(COL_W) + Style.RESET_ALL
        else:
            diff_score_str = ' '*COL_W

        # formata valor operacao
        raw_vo = f"{valor_operacao:,.2f}".center(COL_W)
        if valor_operacao>0:
            val_op_str = Fore.GREEN + raw_vo + Style.RESET_ALL
        elif valor_operacao<0:
            val_op_str = Fore.RED + raw_vo + Style.RESET_ALL
        else:
            val_op_str = raw_vo

        gl_vals = [
            ' '*COL_W,
            val_op_str,
            diff_score_str,
            f"{diff_price:.0f}".center(COL_W),
            f"{brk_price:.0f}".center(COL_W),
            f"{trl_price:.0f}".center(COL_W)
        ]
        for i,(op,gl) in enumerate(zip(op_vals,gl_vals)):
            row = top+2+i
            txt = op.center(COL_W)
            if i==1:
                txt = (Fore.GREEN+txt+Style.RESET_ALL) if entry_decision=='COMPRA' else (Fore.RED+txt+Style.RESET_ALL)
            sys.stdout.write(f"\x1b[{row};{b2}H{txt}")
            sys.stdout.write(f"\x1b[{row};{b3}H{gl}")
        sys.stdout.flush()

    # exibe linha OPERACAO no PLANO DE TRADE
    plan_start = start_row+17
    w = 12
    car_col = 1 + (w+1)*1 + 1
    ban_col = 1 + (w+1)*2 + 1
    ctr_col = 1 + (w+1)*3 + 1
    op_col  = 1 + (w+1)*4 + 1
    stopop_col  = 1 + (w+1)*5 + 1
    metaop_col  = 1 + (w+1)*6 + 1
    stopday_col = 1 + (w+1)*7 + 1
    metaday_col = 1 + (w+1)*8 + 1
    oper_row = plan_start + 5

    # define valores
    if open_position:
        diff_price = (current_price-entry_price) if entry_decision=='COMPRA' else (entry_price-current_price)
        val_op = diff_price * VALOR_PONTO_INDICE
        new_car = plan_data['carteira'] + val_op
        ban = plan_data['banca']
        op_num = operations_executed + 1
    elif last_record:
        new_car = last_record['carteira']
        ban = last_record['banca']
        op_num = last_record.get('op_num',0)
    else:
        new_car = plan_data['carteira']
        ban = plan_data['banca']
        op_num = 0

    # escreve CARTEIRA/BANCA/CONTRATOS/OPERACOES
    raw_car = fmt_brl(new_car)
    car_txt = (
        Fore.GREEN+raw_car+Style.RESET_ALL if ban>0 else
        Fore.RED+raw_car+Style.RESET_ALL if ban<0 else
        raw_car
    )
    raw_ban = fmt_brl(ban)
    ban_txt = (
        Fore.GREEN+raw_ban+Style.RESET_ALL if ban>0 else
        Fore.RED+raw_ban+Style.RESET_ALL if ban<0 else
        raw_ban
    )
    txt_ctr = str(contracts).center(w)
    txt_op  = str(op_num).center(w)

    sys.stdout.write(f"\x1b[{oper_row};{car_col}H{car_txt}")
    sys.stdout.write(f"\x1b[{oper_row};{ban_col}H{ban_txt}")
    sys.stdout.write(f"\x1b[{oper_row};{ctr_col}H{txt_ctr}")
    sys.stdout.write(f"\x1b[{oper_row};{op_col}H{txt_op}")

    # limpa STOP OP e META OP
    sys.stdout.write(f"\x1b[{oper_row};{stopop_col}H{'':^{w}}")
    sys.stdout.write(f"\x1b[{oper_row};{metaop_col}H{'':^{w}}")
    # escreve P&L operação em STOP OP ou META OP
    if open_position and val_op<0:
        sys.stdout.write(f"\x1b[{oper_row};{stopop_col}H" + Fore.RED + fmt_brl(val_op) + Style.RESET_ALL)
    elif open_position and val_op>0:
        sys.stdout.write(f"\x1b[{oper_row};{metaop_col}H" + Fore.GREEN + fmt_brl(val_op) + Style.RESET_ALL)

    # limpa STOP DIARIO e META DIARIA
    sys.stdout.write(f"\x1b[{oper_row};{stopday_col}H{'':^{w}}")
    sys.stdout.write(f"\x1b[{oper_row};{metaday_col}H{'':^{w}}")
    # escreve P&L diário em STOP DIARIO ou META DIARIA
    if daily_pnl<0:
        sys.stdout.write(f"\x1b[{oper_row};{stopday_col}H" + Fore.RED + fmt_brl(daily_pnl) + Style.RESET_ALL)
    elif daily_pnl>0:
        sys.stdout.write(f"\x1b[{oper_row};{metaday_col}H" + Fore.GREEN + fmt_brl(daily_pnl) + Style.RESET_ALL)

    sys.stdout.flush()
